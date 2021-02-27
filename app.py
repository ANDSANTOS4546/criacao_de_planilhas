from openpyxl import Workbook


def opcao_usario(msg):
    while True:
        opcao = str(input(msg)).strip().upper()[0]
        if "S" != opcao != "N":
            continue
        else:
            return opcao


def criar_paginas():
    while True:
        nome_pagina = str(input('Nome da página: ')).strip().title()
        wb.create_sheet(nome_pagina)
    
        adicionar_pagina = opcao_usario('Quer adicionar outra página? [S/N]: ')
        if adicionar_pagina == "N":
            break


def selecionar_pagina():
    escolher_pagina = str(input('Selecio página deseja manipular: ')).strip().title()
    pagina_escolhida = wb[escolher_pagina]
    return pagina_escolhida


def salvar_arquivo():
    nome_arquivo = str(input("Qual nome deseja salvar seu arquivo: ")).capitalize()
    wb.save(nome_arquivo + ".xlsx")


wb = Workbook()
ws = wb.active

# Criação de página
criar_paginas()

del wb['Sheet']  # Deletar a pagina "Sheet"
print(wb.sheetnames)

# Escolher pagina para manipular
selecionar_pagina()

# Insrerir coluna




# Salvar arquivo
salvar_arquivo()
